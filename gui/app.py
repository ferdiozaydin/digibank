from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file
import csv
import requests
import os
import io
import random
from datetime import datetime, timedelta

app = Flask(__name__)
app.secret_key = 'digibank_secret'

# Configuration
API_URL = os.environ.get('BACKEND_URL', 'http://localhost:8080')
VALID_USERNAME = os.environ.get('DIGIBANK_USERNAME', 'admin')
VALID_PASSWORD = os.environ.get('DIGIBANK_PASSWORD', 'admin')
VALID_TOTP = os.environ.get('DIGIBANK_TOTP', '000000')

# Mock data for UI demos
FAKE_CARDS = [
    {
        "id": "DIGI-001",
        "name": "DigiBank Black",
        "brand": "VISA",
        "limit": 50000,
        "available": 38250,
        "status": "Aktif",
        "lastTx": "Trendyol 1.250,00₺",
    },
    {
        "id": "DIGI-002",
        "name": "DigiBank Seyahat",
        "brand": "Mastercard",
        "limit": 20000,
        "available": 7800,
        "status": "Temassız Kapalı",
        "lastTx": "İGA Lounge 450,00₺",
    },
    {
        "id": "DIGI-003",
        "name": "DigiBank Sanal",
        "brand": "VISA",
        "limit": 5000,
        "available": 4900,
        "status": "Aktif",
        "lastTx": "App Store 89,99₺",
    },
]

FAKE_CARD_TX = [
    {"desc": "Market Harcaması", "amount": -320.75, "date": "12.01 14:21"},
    {"desc": "QR Para Çekme", "amount": -1500.0, "date": "12.01 10:04"},
    {"desc": "Kripto Satın Alma", "amount": -250.0, "date": "11.30 20:10"},
    {"desc": "Para İadesi", "amount": 420.0, "date": "11.30 08:15"},
]

DEFAULT_PREFERENCES = {
    "biometric": True,
    "push": True,
    "spendAlerts": True,
    "travelMode": False,
    "darkMode": True,
}

DEFAULT_METRICS = {
    "apiLatencyMs": 120,
    "paymentSuccessRate": 98.5,
    "sensorAlertsActive": 3,
    "cryptoUsagePct": 42,
}

# Basit zaman serisi demo verisi (dashboard'a ek performans sayfası için)
DEFAULT_METRIC_TIMELINE = [
    {"label": "09:00", "latency": 148, "success": 97.8, "throughput": 162, "queue": 3},
    {"label": "10:00", "latency": 131, "success": 98.2, "throughput": 175, "queue": 2},
    {"label": "11:00", "latency": 118, "success": 98.6, "throughput": 184, "queue": 2},
    {"label": "12:00", "latency": 126, "success": 98.0, "throughput": 191, "queue": 3},
    {"label": "13:00", "latency": 112, "success": 98.9, "throughput": 205, "queue": 1},
    {"label": "14:00", "latency": 107, "success": 99.1, "throughput": 212, "queue": 1},
    {"label": "15:00", "latency": 115, "success": 98.7, "throughput": 208, "queue": 2},
]

FX_RATES = {
    "USD": {"TRY": 31.2, "EUR": 0.92, "GBP": 0.79},
    "EUR": {"TRY": 34.1, "USD": 1.08, "GBP": 0.86},
    "GBP": {"TRY": 39.5, "USD": 1.26, "EUR": 1.17},
    "TRY": {"USD": 0.032, "EUR": 0.029, "GBP": 0.025},
}

DEMO_ACCOUNTS = [
    {"iban": "TR12 0001 0000 1234 5678 0001", "name": "Vadesiz TL", "balance": 18500.75},
    {"iban": "TR34 0001 0000 1234 5678 0002", "name": "USD Hesabı", "balance": 5200.00},
    {"iban": "TR56 0001 0000 1234 5678 0003", "name": "EUR Hesabı", "balance": 3100.50},
]


def _normalize_user(raw: dict) -> dict:
    """Temel kullanici sozluk yapisini sabitle ve bos alanlari doldur."""
    return {
        "username": session.get("username") or raw.get("username", "Bağlantı yok"),
        "fiatBalance": raw.get("fiatBalance", 0),
        "cryptoBalance": raw.get("cryptoBalance", 0),
        "transactions": raw.get("transactions", []),
    }


def _is_authenticated() -> bool:
    return bool(session.get('logged_in'))


def _auth_headers():
    token = session.get('api_token')
    if token:
        return {"Authorization": f"Bearer {token}"}
    return {}


def _handle_unauthorized(resp, next_path: str = "/"):
    """API 401 dönerse (örn: backend yeniden başladı), GUI oturumunu sıfırla ve login'e yönlendir."""
    try:
        status = getattr(resp, "status_code", None)
    except Exception:
        status = None
    if status == 401:
        session.clear()
        flash('Oturum doğrulanamadı (token geçersiz/eksik). Lütfen tekrar giriş yapın.', 'error')
        return redirect(url_for('login', next=next_path))
    return None


def _get_preferences() -> dict:
    prefs = session.get('preferences') or {}
    merged = DEFAULT_PREFERENCES.copy()
    merged.update({k: bool(v) for k, v in prefs.items()})
    return merged


def _append_card_event(card_id: str, action: str):
    events = session.get('card_events') or []
    events.insert(0, {
        "cardId": card_id or "Bilinmiyor",
        "action": action or "İşlem",
        "ts": datetime.now().strftime("%d.%m %H:%M")
    })
    session['card_events'] = events[:12]


def _get_fx_rate(src: str, dst: str):
    if not src or not dst:
        return None
    if src == dst:
        return 1.0
    return FX_RATES.get(src, {}).get(dst)


def _daily_rng(salt: str = "") -> random.Random:
    """Günlük değişen ama sayfa yenilemede stabil demo verisi için RNG."""
    day_seed = datetime.now().strftime("%Y%m%d")
    user_seed = session.get("username") or "guest"
    return random.Random(f"{day_seed}:{user_seed}:{salt}")


def _is_valid_manual_tx(tx) -> bool:
    """Transactions sayfasındaki tablo sadece manuel kayıt alanlarını gösterir.

    Bu yüzden eski model (description/userId) kayıtlar veya tamamen boş placeholder
    kayıtlar tabloda görünmesin.
    """
    if not isinstance(tx, dict):
        return True

    tx_id = tx.get("id")
    if tx_id is None or str(tx_id).strip() == "":
        return False

    full_name = str(tx.get("fullName") or "").strip()
    bank_code = str(tx.get("bankCode") or "").strip()
    address = str(tx.get("address") or "").strip()
    record_date = str(tx.get("recordDate") or "").strip()
    amount = tx.get("amount")

    if not (full_name and bank_code and address and record_date):
        return False

    if amount is None:
        return False
    if isinstance(amount, str) and amount.strip() == "":
        return False

    return True


def _try_parse_date(date_str: str):
    if not date_str:
        return None
    val = str(date_str).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d", "%d.%m", "%d/%m"):
        try:
            dt = datetime.strptime(val, fmt)
            if fmt in ("%d.%m", "%d/%m"):
                dt = dt.replace(year=datetime.now().year)
            return dt
        except ValueError:
            continue
    return None


def _bill_type_meta(raw_type: str) -> dict:
    t = (raw_type or "").strip().upper()
    mapping = {
        "TAX": ("Vergi Borcu", "Gelir İdaresi"),
        "TRAFFIC_FINE": ("Trafik Cezası", "Emniyet Genel Müdürlüğü"),
        "UTILITY": ("Kamu Hizmeti", "Belediye / Kamu"),
        "MTV": ("Motorlu Taşıtlar Vergisi", "Gelir İdaresi"),
        "SGK": ("SGK Prim Borcu", "SGK"),
        "PROPERTY_TAX": ("Emlak Vergisi", "Belediye"),
        "PASSPORT": ("Pasaport Harcı", "Nüfus ve Vatandaşlık"),
    }
    label, institution = mapping.get(t, (raw_type or "Devlet Ödemesi", "Kamu"))
    return {"label": label, "institution": institution}


def _generate_demo_bills(now: datetime):
    rng = _daily_rng("bills")
    templates = [
        ("TAX", 825.40, 6),
        ("TRAFFIC_FINE", 1420.00, 2),
        ("UTILITY", 356.90, 11),
        ("SGK", 2150.75, 18),
        ("MTV", 2975.00, 4),
        ("PROPERTY_TAX", 980.30, 14),
        ("PASSPORT", 215.00, 30),
    ]

    rng.shuffle(templates)
    count = rng.randint(4, 6)
    picked = templates[:count]

    bills = []
    for idx, (bill_type, base_amount, due_in_days) in enumerate(picked, start=1):
        jitter = rng.uniform(-0.08, 0.12)
        amount = round(base_amount * (1 + jitter), 2)
        due = (now + timedelta(days=due_in_days)).date().isoformat()
        bill_id = f"GOV-{now.strftime('%Y%m')}-{rng.randint(1000, 9999)}-{idx:02d}"
        # Bazılarını ödenmiş göster
        is_paid = rng.random() < 0.25
        bills.append({
            "billingId": bill_id,
            "type": bill_type,
            "amount": amount,
            "date": due,
            "isPaid": bool(is_paid),
        })

    return bills


def _decorate_bills(bills: list, now: datetime) -> list:
    out = []
    for b in (bills or []):
        billing_id = (b.get("billingId") if isinstance(b, dict) else getattr(b, "billingId", None))
        raw_type = (b.get("type") if isinstance(b, dict) else getattr(b, "type", None))
        amount_raw = (b.get("amount") if isinstance(b, dict) else getattr(b, "amount", 0))
        date_raw = (b.get("date") if isinstance(b, dict) else getattr(b, "date", None))
        is_paid = (b.get("isPaid") if isinstance(b, dict) else getattr(b, "isPaid", False))

        try:
            amount = float(amount_raw)
        except Exception:
            amount = 0.0

        due_dt = _try_parse_date(date_raw)
        if due_dt:
            days_left = (due_dt.date() - now.date()).days
            due_text = due_dt.strftime("%d.%m.%Y")
        else:
            days_left = None
            due_text = "-"

        if days_left is None:
            days_text = "-"
        elif days_left < 0:
            days_text = f"gecikti ({abs(days_left)}g)"
        else:
            days_text = f"{days_left} gün"

        meta = _bill_type_meta(raw_type)

        if bool(is_paid):
            urgency = "ok"
            status_text = "Ödeme tamam"
        else:
            if days_left is None:
                urgency = "warn"
                status_text = "Ödeme bekliyor"
            elif days_left < 0:
                urgency = "error"
                status_text = "Gecikti"
            elif days_left <= 3:
                urgency = "warn"
                status_text = "Son günler"
            else:
                urgency = "ghost"
                status_text = "Ödeme bekliyor"

        out.append({
            "billingId": billing_id or "-",
            "rawType": raw_type or "-",
            "typeLabel": meta["label"],
            "institution": meta["institution"],
            "amount": amount,
            "date": date_raw or "-",
            "dueText": due_text,
            "daysLeft": days_left,
            "daysText": days_text,
            "isPaid": bool(is_paid),
            "urgency": urgency,
            "statusText": status_text,
        })

    def _sort_key(row):
        unpaid_rank = 0 if not row["isPaid"] else 1
        if row["daysLeft"] is None:
            return (unpaid_rank, 9999, row["billingId"])
        return (unpaid_rank, row["daysLeft"], row["billingId"])

    out.sort(key=_sort_key)
    return out


def _bill_summary(decorated: list) -> dict:
    pending = [b for b in decorated if not b.get("isPaid")]
    paid = [b for b in decorated if b.get("isPaid")]
    pending_total = round(sum((b.get("amount") or 0) for b in pending), 2)
    paid_total = round(sum((b.get("amount") or 0) for b in paid), 2)
    soonest = None
    soonest_days = None
    for b in pending:
        d = b.get("daysLeft")
        if d is None:
            continue
        if soonest_days is None or d < soonest_days:
            soonest_days = d
            soonest = b.get("dueText")
    return {
        "pendingCount": len(pending),
        "pendingTotal": pending_total,
        "paidCount": len(paid),
        "paidTotal": paid_total,
        "soonestDueText": soonest or "-",
    }


def _generate_home_dashboard(now: datetime) -> dict:
    rng = _daily_rng("home")
    indoor_temp = round(rng.uniform(19.5, 24.5), 1)
    humidity = rng.randint(32, 58)
    active_devices = rng.randint(6, 14)
    energy_today = round(rng.uniform(5.8, 18.4), 1)
    wifi_quality = rng.randint(78, 99)

    systems = [
        {
            "name": "Güvenlik Sistemi",
            "detail": rng.choice(["Alarm kurulu", "Ev modu", "Gece modu"]),
            "status": "ok" if rng.random() < 0.82 else "warn",
        },
        {
            "name": "Yangın / Duman Sensörleri",
            "detail": rng.choice(["Normal", "Normal", "Pil düşük: Mutfak"]),
            "status": "ok" if rng.random() < 0.75 else "warn",
        },
        {
            "name": "Akıllı Aydınlatma",
            "detail": f"{rng.randint(1, 6)} oda açık",
            "status": "ok",
        },
        {
            "name": "Isıtma / Soğutma (HVAC)",
            "detail": f"Hedef {round(indoor_temp + rng.uniform(-1.0, 1.0), 1)}°C",
            "status": "ok" if rng.random() < 0.85 else "warn",
        },
        {
            "name": "Su Kaçağı Sensörleri",
            "detail": rng.choice(["Normal", "Normal", "Son kontrol: 2 dk önce"]),
            "status": "ok",
        },
        {
            "name": "Wi‑Fi / Ağ",
            "detail": f"Sinyal {wifi_quality}%",
            "status": "ok" if wifi_quality >= 85 else "warn",
        },
    ]

    device_templates = [
        ("LIGHT-LIVINGROOM", "Salon Lambası", "Salon"),
        ("LIGHT-KITCHEN", "Mutfak Lambası", "Mutfak"),
        ("PLUG-COFFEE", "Kahve Prizi", "Mutfak"),
        ("LOCK-ENTRANCE", "Giriş Kilidi", "Antre"),
        ("CAM-DOOR", "Kapı Kamerası", "Giriş"),
        ("THERMO-SALON", "Termostat", "Salon"),
        ("SPRINKLER", "Bahçe Sulama", "Bahçe"),
    ]
    rng.shuffle(device_templates)
    devices = []
    for dev_id, name, room in device_templates[: rng.randint(5, 7)]:
        state_on = rng.random() < 0.55
        watts = rng.randint(3, 140) if state_on and ("LIGHT" in dev_id or "PLUG" in dev_id) else rng.randint(0, 12)
        devices.append({
            "id": dev_id,
            "name": name,
            "room": room,
            "state": "Açık" if state_on else "Kapalı",
            "watts": watts,
            "updated": (now - timedelta(minutes=rng.randint(1, 55))).strftime("%H:%M"),
            "pill": "live" if state_on else "ghost",
        })

    return {
        "kpis": {
            "indoorTemp": indoor_temp,
            "humidity": humidity,
            "activeDevices": active_devices,
            "energyToday": energy_today,
        },
        "systems": systems,
        "devices": devices,
        "updatedAt": now.strftime("%d.%m %H:%M"),
    }


@app.before_request
def require_login():
    allowed_endpoints = {'login', 'static'}
    if request.endpoint in allowed_endpoints or request.endpoint is None:
        return

    if not _is_authenticated():
        flash('Lütfen giriş yapınız.', 'error')
        return redirect(url_for('login', next=request.path))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        totp = request.form.get('totp', '').strip() or VALID_TOTP

        if username == VALID_USERNAME and password == VALID_PASSWORD:
            # Request backend token with MFA code
            try:
                resp = requests.post(
                    f"{API_URL}/api/login",
                    json={"username": username, "password": password, "totp": totp},
                    timeout=5,
                )
                data = resp.json() if resp.status_code == 200 else {}
                token = data.get('token')
                if not token:
                    raise ValueError('API token alınamadı')
                session['api_token'] = token
            except Exception as e:
                flash(f'API oturum açılamadı: {e}', 'error')
                return render_template(
                    'login.html',
                    user=_normalize_user({}),
                    default_username=VALID_USERNAME,
                    default_password=VALID_PASSWORD,
                    default_totp=VALID_TOTP,
                )

            session['logged_in'] = True
            session['username'] = username
            flash('Giriş başarılı.', 'success')
            next_url = request.args.get('next')
            return redirect(next_url or url_for('dashboard'))

        flash('Geçersiz kullanıcı adı veya şifre.', 'error')

    return render_template(
        'login.html',
        user=_normalize_user({}),
        default_username=VALID_USERNAME,
        default_password=VALID_PASSWORD,
        default_totp=VALID_TOTP,
    )


@app.route('/logout')
def logout():
    session.clear()
    flash('Oturum kapatıldı.', 'success')
    return redirect(url_for('login'))

@app.route('/')
def dashboard():
    # Fetch User Data from Java Backend
    try:
        resp = requests.get(f"{API_URL}/api/user", timeout=3, headers=_auth_headers())
        user_data = resp.json() if resp.status_code == 200 else {}
    except Exception:
        user_data = {}

    try:
        metrics_resp = requests.get(f"{API_URL}/api/metrics", timeout=3, headers=_auth_headers())
        metrics = metrics_resp.json() if metrics_resp.status_code == 200 else {}
    except Exception:
        metrics = {}

    merged_metrics = {**DEFAULT_METRICS, **{k: v for k, v in (metrics or {}).items() if v is not None}}

    return render_template(
        'dashboard.html',
        user=_normalize_user(user_data),
        metrics=merged_metrics,
        metrics_live=bool(metrics),
    )


@app.route('/performance')
def performance():
    # Kullanıcıyı çek (başlıkta isim gözüksün)
    try:
        user_resp = requests.get(f"{API_URL}/api/user", timeout=3, headers=_auth_headers())
        user_data = user_resp.json() if user_resp.status_code == 200 else {}
    except Exception:
        user_data = {}

    # Canlı metrikleri dene, yoksa demo verisine düş
    try:
        metrics_resp = requests.get(f"{API_URL}/api/metrics", timeout=3, headers=_auth_headers())
        metrics = metrics_resp.json() if metrics_resp.status_code == 200 else {}
    except Exception:
        metrics = {}

    merged_metrics = {**DEFAULT_METRICS, **{k: v for k, v in (metrics or {}).items() if v is not None}}

    # Zaman serisi verisi: canlı metrik varsa son noktayı ekle
    timeline = [dict(row) for row in DEFAULT_METRIC_TIMELINE]
    if merged_metrics:
        timeline.append({
            "label": datetime.now().strftime("%H:%M"),
            "latency": merged_metrics.get("apiLatencyMs", DEFAULT_METRICS["apiLatencyMs"]),
            "success": merged_metrics.get("paymentSuccessRate", DEFAULT_METRICS["paymentSuccessRate"]),
            "throughput": merged_metrics.get("transactionsPerMin", 210),
            "queue": merged_metrics.get("queueDepth", 2),
        })

    return render_template(
        'performance.html',
        user=_normalize_user(user_data),
        metrics=merged_metrics,
        metrics_live=bool(metrics),
        timeline=timeline,
    )


@app.route('/cards', methods=['GET', 'POST'])
def cards():
    if request.method == 'POST':
        action = request.form.get('action', '').strip() or 'İşlem'
        card_id = request.form.get('cardId', '').strip() or 'Kart'
        _append_card_event(card_id, action)
        flash(f"{card_id} için '{action}' isteği oluşturuldu (simülasyon).", 'success')
        return redirect(url_for('cards'))

    try:
        resp = requests.get(f"{API_URL}/api/user", timeout=3, headers=_auth_headers())
        user_data = resp.json() if resp.status_code == 200 else {}
    except Exception:
        user_data = {}

    return render_template(
        'cards.html',
        user=_normalize_user(user_data),
        cards=FAKE_CARDS,
        tx=FAKE_CARD_TX,
        events=session.get('card_events', []),
    )


@app.route('/home-control')
def home_control():
    # We can reuse user fetch for greeting
    try:
        resp = requests.get(f"{API_URL}/api/user", timeout=3, headers=_auth_headers())
        user_data = resp.json() if resp.status_code == 200 else {}
    except Exception:
        user_data = {}
    home_dash = _generate_home_dashboard(datetime.now())
    return render_template('home_control.html', user=_normalize_user(user_data), home_dash=home_dash)


@app.route('/home/toggle', methods=['POST'])
def home_toggle():
    device_id = request.form.get('deviceId', '').strip()
    state = request.form.get('state', 'off')
    on = state.lower() in ('on', 'true', '1')
    try:
        resp = requests.post(
            f"{API_URL}/api/home/toggle",
            json={"deviceId": device_id, "on": str(on).lower()},
            timeout=5,
            headers=_auth_headers(),
        )
        if resp.status_code == 200:
            flash(f"{device_id} için durum güncellendi", 'success')
        else:
            flash(resp.json().get('hata', 'Cihaz kontrolü başarısız'), 'error')
    except Exception:
        flash('Cihaz kontrolü başarısız: bağlantı hatası', 'error')
    return redirect(url_for('home_control'))


@app.route('/home/thermostat', methods=['POST'])
def home_thermostat():
    zone = request.form.get('zone', '').strip()
    target = request.form.get('target', '').strip()
    try:
        resp = requests.post(
            f"{API_URL}/api/home/thermostat",
            json={"zone": zone, "target": target},
            timeout=5,
            headers=_auth_headers(),
        )
        if resp.status_code == 200:
            flash(f"{zone} için hedef sıcaklık ayarlandı", 'success')
        else:
            flash(resp.json().get('hata', 'Termostat ayarı başarısız'), 'error')
    except Exception:
        flash('Termostat ayarı başarısız: bağlantı hatası', 'error')
    return redirect(url_for('home_control'))

@app.route('/smart-gov')
def smart_gov():
    now = datetime.now()
    try:
        user_resp = requests.get(f"{API_URL}/api/user", timeout=3, headers=_auth_headers())
        user_data = user_resp.json() if user_resp.status_code == 200 else {}

        bills_resp = requests.get(f"{API_URL}/api/bills", timeout=3, headers=_auth_headers())
        bills_data = bills_resp.json() if bills_resp.status_code == 200 else []
    except Exception as e:
        print(e)
        user_data = {}
        bills_data = []

    if not bills_data:
        bills_data = _generate_demo_bills(now)

    decorated = _decorate_bills(bills_data, now)
    summary = _bill_summary(decorated)

    return render_template(
        'smart_gov.html',
        bills=decorated,
        bill_summary=summary,
        bills_updated_at=now.strftime("%d.%m %H:%M"),
        user=_normalize_user(user_data),
    )


@app.route('/fx', methods=['GET', 'POST'])
def fx_swap():
    try:
        resp = requests.get(f"{API_URL}/api/user", timeout=3, headers=_auth_headers())
        user_data = resp.json() if resp.status_code == 200 else {}
    except Exception:
        user_data = {}

    swap_result = None
    if request.method == 'POST':
        src = request.form.get('fromCurrency', 'USD').strip().upper()
        dst = request.form.get('toCurrency', 'TRY').strip().upper()
        try:
            amount = float(request.form.get('amount', 0) or 0)
        except ValueError:
            amount = 0
        rate = _get_fx_rate(src, dst)
        if rate:
            converted = round(amount * rate, 2)
            swap_result = {
                'src': src,
                'dst': dst,
                'amount': amount,
                'rate': rate,
                'converted': converted,
            }
            flash(f"{amount} {src} -> {converted} {dst} işlemi hazır (demo).", 'success')
        else:
            flash('Kur bulunamadı, lütfen farklı bir çift seçin.', 'error')

    return render_template(
        'fx_swap.html',
        user=_normalize_user(user_data),
        rates=FX_RATES,
        swap_result=swap_result,
    )


@app.route('/transfer', methods=['GET', 'POST'])
def transfer():
    try:
        resp = requests.get(f"{API_URL}/api/user", timeout=3, headers=_auth_headers())
        user_data = resp.json() if resp.status_code == 200 else {}
    except Exception:
        user_data = {}

    if request.method == 'POST':
        to_name = request.form.get('toName', 'Alıcı').strip() or 'Alıcı'
        iban = request.form.get('iban', '').strip()
        amount = request.form.get('amount', '').strip()
        desc = request.form.get('description', '').strip()
        try:
            resp = requests.post(
                f"{API_URL}/api/transfer",
                json={"toName": to_name, "iban": iban, "amount": str(amount), "description": desc},
                timeout=5,
                headers=_auth_headers(),
            )
            if resp.status_code == 201:
                flash(f"{to_name} adına {amount} TL transfer kaydı oluşturuldu.", 'success')
            else:
                msg = resp.json().get('hata', 'Transfer başarısız') if resp.content else 'Transfer başarısız'
                flash(msg, 'error')
        except Exception:
            flash('Transfer başarısız: bağlantı hatası', 'error')
        return redirect(url_for('transfer'))

    return render_template(
        'transfer.html',
        user=_normalize_user(user_data),
        accounts=DEMO_ACCOUNTS,
    )

@app.route('/pay/<bill_id>', methods=['POST'])
def pay_bill(bill_id):
    pay_mode = request.form.get('payMode', 'FIAT')
    try:
        resp = requests.post(
            f"{API_URL}/api/pay",
            json={"billId": bill_id, "payMode": pay_mode},
            timeout=5,
            headers=_auth_headers(),
        )
        try:
            data = resp.json()
        except ValueError:
            data = {}

        if resp.status_code == 200 and data.get('durum') == 'BASARILI':
            flash(f"{bill_id} numaralı fatura ödendi", 'success')
        else:
            flash(data.get('hata') or 'Ödeme başarısız: API hatası', 'error')
    except Exception:
         flash('Ödeme başarısız: Bağlantı hatası', 'error')
    
    return redirect(url_for('smart_gov'))

@app.route('/export', methods=['POST'])
def export_data():
    try:
        resp = requests.post(f"{API_URL}/api/export", timeout=5, headers=_auth_headers())
        if resp.status_code == 200:
            data = resp.json()
            flash(f"Dışarı aktarma başarılı! Oluşan dosya: {data.get('dosya')}", "success")
        else:
            flash("Dışarı aktarma başarısız.", "error")
    except Exception as e:
        flash("Dışarı aktarma hatası: Bağlantı sorunu", "error")
    
    return redirect(url_for('dashboard'))


@app.route('/export/download')
def export_download():
    """Basit CSV çıktı ile istemci tarafında indirilebilir içerik üretir."""
    try:
        resp = requests.get(f"{API_URL}/api/user", timeout=3, headers=_auth_headers())
        user_data = resp.json() if resp.status_code == 200 else {}
    except Exception:
        user_data = {}

    user = _normalize_user(user_data)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    csv_lines = [
        "exported_at,username,fiat_balance,crypto_balance",
        f"{now},{user['username']},{user['fiatBalance']},{user['cryptoBalance']}",
    ]
    csv_content = "\n".join(csv_lines)
    file_buf = io.BytesIO(csv_content.encode("utf-8"))
    filename = f"digibank_user_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return send_file(file_buf, as_attachment=True, download_name=filename, mimetype='text/csv')


@app.route('/settings', methods=['GET', 'POST'])
def settings():
    prefs = _get_preferences()

    if request.method == 'POST':
        prefs['biometric'] = bool(request.form.get('biometric'))
        prefs['push'] = bool(request.form.get('push'))
        prefs['spendAlerts'] = bool(request.form.get('spendAlerts'))
        prefs['travelMode'] = bool(request.form.get('travelMode'))
        prefs['darkMode'] = bool(request.form.get('darkMode'))
        session['preferences'] = prefs

        logs = session.get('settings_logs') or []
        logs.insert(0, {
            "ts": datetime.now().strftime("%d.%m %H:%M"),
            "note": f"Ayarlar güncellendi: push={'açık' if prefs['push'] else 'kapalı'}, travel={'açık' if prefs['travelMode'] else 'kapalı'}",
        })
        session['settings_logs'] = logs[:10]

        flash('Ayarlar kaydedildi (simülasyon).', 'success')
        return redirect(url_for('settings'))

    try:
        resp = requests.get(f"{API_URL}/api/user", timeout=3, headers=_auth_headers())
        user_data = resp.json() if resp.status_code == 200 else {}
    except Exception:
        user_data = {}

    return render_template(
        'settings.html',
        user=_normalize_user(user_data),
        preferences=prefs,
        logs=session.get('settings_logs', []),
    )


@app.route('/users', methods=['GET', 'POST'])
def users():
    if request.method == 'POST':
        action = (request.form.get('action') or 'create').strip().lower()

        if action == 'create':
            new_username = request.form.get('username', '').strip()
            new_password = request.form.get('password', '').strip()
            role = request.form.get('role', 'RESIDENT').strip().upper() or 'RESIDENT'
            if not new_username or not new_password:
                flash('Kullanıcı adı ve şifre gerekli.', 'error')
            else:
                try:
                    resp = requests.post(
                        f"{API_URL}/api/users/register",
                        json={"username": new_username, "password": new_password, "role": role},
                        headers=_auth_headers(),
                        timeout=5,
                    )
                    if resp.status_code == 201:
                        flash(f"{new_username} eklendi.", 'success')
                    else:
                        msg = resp.json().get('hata', 'Kayıt başarısız') if resp.content else 'Kayıt başarısız'
                        flash(msg, 'error')
                except Exception:
                    flash('Kayıt başarısız: bağlantı hatası', 'error')
            return redirect(url_for('users'))

        if action == 'update':
            username = request.form.get('edit_username', '').strip()
            role = request.form.get('edit_role', '').strip().upper()
            fiat = (request.form.get('edit_fiat') or '').strip()
            crypto = (request.form.get('edit_crypto') or '').strip()
            try:
                resp = requests.put(
                    f"{API_URL}/api/users/item",
                    params={"username": username},
                    json={"role": role, "fiatBalance": fiat, "cryptoBalance": crypto},
                    headers=_auth_headers(),
                    timeout=5,
                )
                if resp.status_code == 200:
                    flash(f"{username} güncellendi.", 'success')
                else:
                    msg = resp.json().get('hata', 'Güncelleme başarısız') if resp.content else 'Güncelleme başarısız'
                    flash(msg, 'error')
            except Exception:
                flash('Güncelleme başarısız: bağlantı hatası', 'error')
            return redirect(url_for('users'))

        if action == 'delete':
            username = request.form.get('delete_username', '').strip()
            try:
                resp = requests.delete(
                    f"{API_URL}/api/users/item",
                    params={"username": username},
                    headers=_auth_headers(),
                    timeout=5,
                )
                if resp.status_code == 200:
                    flash(f"{username} silindi.", 'success')
                else:
                    msg = resp.json().get('hata', 'Silme başarısız') if resp.content else 'Silme başarısız'
                    flash(msg, 'error')
            except Exception:
                flash('Silme başarısız: bağlantı hatası', 'error')
            return redirect(url_for('users'))

        flash('Bilinmeyen işlem.', 'error')
        return redirect(url_for('users'))

    try:
        q = (request.args.get('q') or '').strip()
        if q:
            list_resp = requests.get(
                f"{API_URL}/api/users/search",
                params={"q": q},
                headers=_auth_headers(),
                timeout=5,
            )
        else:
            list_resp = requests.get(f"{API_URL}/api/users", headers=_auth_headers(), timeout=5)
        users_data = list_resp.json() if list_resp.status_code == 200 else []
    except Exception:
        users_data = []

    try:
        resp = requests.get(f"{API_URL}/api/user", timeout=3, headers=_auth_headers())
        user_data = resp.json() if resp.status_code == 200 else {}
    except Exception:
        user_data = {}

    return render_template('users.html', user=_normalize_user(user_data), users=users_data, q=(request.args.get('q') or '').strip())


@app.route('/transactions', methods=['GET', 'POST'])
def transactions():
    if request.method == 'POST':
        action = (request.form.get('action') or 'create').strip().lower()

        if action == 'create':
            full_name = (request.form.get('fullName') or '').strip()
            amount = (request.form.get('amount') or '').strip()
            bank_code = (request.form.get('bankCode') or '').strip()
            address = (request.form.get('address') or '').strip()
            record_date = (request.form.get('recordDate') or '').strip()
            try:
                resp = requests.post(
                    f"{API_URL}/api/transactions/create",
                    json={
                        "fullName": full_name,
                        "amount": str(amount),
                        "bankCode": bank_code,
                        "address": address,
                        "recordDate": record_date,
                    },
                    headers=_auth_headers(),
                    timeout=5,
                )
                redir = _handle_unauthorized(resp, next_path=request.path)
                if redir:
                    return redir
                if resp.status_code == 201:
                    flash('İşlem kaydı eklendi.', 'success')
                else:
                    msg = resp.json().get('hata', 'Kayıt başarısız') if resp.content else 'Kayıt başarısız'
                    flash(msg, 'error')
            except Exception:
                flash('Kayıt başarısız: bağlantı hatası', 'error')
            return redirect(url_for('transactions'))

        if action == 'update':
            tx_id = (request.form.get('edit_id') or '').strip()
            full_name = (request.form.get('edit_fullName') or '').strip()
            amount = (request.form.get('edit_amount') or '').strip()
            bank_code = (request.form.get('edit_bankCode') or '').strip()
            address = (request.form.get('edit_address') or '').strip()
            record_date = (request.form.get('edit_recordDate') or '').strip()
            try:
                resp = requests.put(
                    f"{API_URL}/api/transactions/item",
                    params={"id": tx_id},
                    json={
                        "fullName": full_name,
                        "amount": str(amount),
                        "bankCode": bank_code,
                        "address": address,
                        "recordDate": record_date,
                    },
                    headers=_auth_headers(),
                    timeout=5,
                )
                redir = _handle_unauthorized(resp, next_path=request.path)
                if redir:
                    return redir
                if resp.status_code == 200:
                    flash('İşlem güncellendi.', 'success')
                else:
                    msg = resp.json().get('hata', 'Güncelleme başarısız') if resp.content else 'Güncelleme başarısız'
                    flash(msg, 'error')
            except Exception:
                flash('Güncelleme başarısız: bağlantı hatası', 'error')
            return redirect(url_for('transactions'))

        if action == 'delete':
            tx_id = (request.form.get('delete_id') or '').strip()
            try:
                resp = requests.delete(
                    f"{API_URL}/api/transactions/item",
                    params={"id": tx_id},
                    headers=_auth_headers(),
                    timeout=5,
                )
                redir = _handle_unauthorized(resp, next_path=request.path)
                if redir:
                    return redir
                if resp.status_code == 200:
                    flash('İşlem silindi.', 'success')
                else:
                    msg = resp.json().get('hata', 'Silme başarısız') if resp.content else 'Silme başarısız'
                    flash(msg, 'error')
            except Exception:
                flash('Silme başarısız: bağlantı hatası', 'error')
            return redirect(url_for('transactions'))

        flash('Bilinmeyen işlem.', 'error')
        return redirect(url_for('transactions'))

    # GET: list/search
    try:
        q = (request.args.get('q') or '').strip()
        if q:
            resp = requests.get(
                f"{API_URL}/api/transactions/search",
                params={"q": q},
                headers=_auth_headers(),
                timeout=5,
            )
        else:
            resp = requests.get(f"{API_URL}/api/transactions", headers=_auth_headers(), timeout=5)
        redir = _handle_unauthorized(resp, next_path=request.path)
        if redir:
            return redir
        txs = resp.json() if resp.status_code == 200 else []
    except Exception:
        txs = []

    try:
        txs = [tx for tx in (txs or []) if _is_valid_manual_tx(tx)]
    except Exception:
        txs = []

    try:
        uresp = requests.get(f"{API_URL}/api/user", timeout=3, headers=_auth_headers())
        user_data = uresp.json() if uresp.status_code == 200 else {}
    except Exception:
        user_data = {}

    bank_codes = [
        "AKBANK",
        "GARANTI",
        "ISBANK",
        "YAPIKREDI",
        "QNB",
        "DENIZBANK",
        "HALKBANK",
        "VAKIFBANK",
        "ZIRAAT",
        "ING",
    ]
    rng = _daily_rng("banks")
    rng.shuffle(bank_codes)

    return render_template(
        'transactions.html',
        user=_normalize_user(user_data),
        txs=txs,
        q=(request.args.get('q') or '').strip(),
        today=datetime.now().date().isoformat(),
        bank_codes=bank_codes,
    )


@app.route('/transactions/download')
def transactions_download():
    """Transactions listesini Excel (XLSX) olarak indirir (q parametresi desteklenir)."""
    q = (request.args.get('q') or '').strip()

    try:
        if q:
            resp = requests.get(
                f"{API_URL}/api/transactions/search",
                params={"q": q},
                headers=_auth_headers(),
                timeout=5,
            )
        else:
            resp = requests.get(f"{API_URL}/api/transactions", headers=_auth_headers(), timeout=5)

        redir = _handle_unauthorized(resp, next_path=request.path)
        if redir:
            return redir

        txs = resp.json() if resp.status_code == 200 else []
    except Exception:
        txs = []

    try:
        txs = [tx for tx in (txs or []) if _is_valid_manual_tx(tx)]
    except Exception:
        txs = []

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "transactions"
    ws.append(["id", "fullName", "amount", "bankCode", "address", "recordDate"])
    for tx in txs:
        ws.append(
            [
                tx.get("id", ""),
                tx.get("fullName", ""),
                tx.get("amount", ""),
                tx.get("bankCode", ""),
                tx.get("address", ""),
                tx.get("recordDate", ""),
            ]
        )

    file_buf = io.BytesIO()
    wb.save(file_buf)
    file_buf.seek(0)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_q = (q[:40] if q else "all")
    filename = f"digibank_transactions_{safe_q}_{stamp}.xlsx"
    return send_file(
        file_buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/transactions/download/pdf')
def transactions_download_pdf():
    """Transactions listesini PDF olarak indirir (q parametresi desteklenir)."""
    q = (request.args.get('q') or '').strip()

    try:
        if q:
            resp = requests.get(
                f"{API_URL}/api/transactions/search",
                params={"q": q},
                headers=_auth_headers(),
                timeout=5,
            )
        else:
            resp = requests.get(f"{API_URL}/api/transactions", headers=_auth_headers(), timeout=5)

        redir = _handle_unauthorized(resp, next_path=request.path)
        if redir:
            return redir

        txs = resp.json() if resp.status_code == 200 else []
    except Exception:
        txs = []

    try:
        txs = [tx for tx in (txs or []) if _is_valid_manual_tx(tx)]
    except Exception:
        txs = []

    def _ascii_tr(s):
        s = "" if s is None else str(s)
        table = str.maketrans({
            "ç": "c",
            "Ç": "C",
            "ğ": "g",
            "Ğ": "G",
            "ı": "i",
            "İ": "I",
            "ö": "o",
            "Ö": "O",
            "ş": "s",
            "Ş": "S",
            "ü": "u",
            "Ü": "U",
        })
        return s.translate(table)

    from fpdf import FPDF

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 14)
    title = f"DigiBank - Islem Listesi ({_ascii_tr(q) if q else 'all'})"
    pdf.cell(0, 10, title, ln=True)

    headers = ["ID", "Ad Soyad", "Para", "Banka", "Adres", "Tarih"]
    widths = [16, 55, 25, 30, 120, 28]

    pdf.set_font("Helvetica", "B", 10)
    for h, w in zip(headers, widths):
        pdf.cell(w, 8, _ascii_tr(h), border=1)
    pdf.ln(8)

    pdf.set_font("Helvetica", "", 9)
    for tx in txs:
        row = [
            tx.get("id", ""),
            tx.get("fullName", ""),
            tx.get("amount", ""),
            tx.get("bankCode", ""),
            tx.get("address", ""),
            tx.get("recordDate", ""),
        ]
        row = [_ascii_tr(v) for v in row]

        # Basit satır: adresi çok uzunsa kes.
        if len(row[4]) > 80:
            row[4] = row[4][:77] + "..."

        for v, w in zip(row, widths):
            pdf.cell(w, 7, v, border=1)
        pdf.ln(7)

    raw_pdf = pdf.output(dest="S")
    if isinstance(raw_pdf, (bytes, bytearray)):
        pdf_bytes = bytes(raw_pdf)
    else:
        pdf_bytes = str(raw_pdf).encode("latin-1", errors="replace")
    file_buf = io.BytesIO(pdf_bytes)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_q = (q[:40] if q else "all")
    filename = f"digibank_transactions_{safe_q}_{stamp}.pdf"
    return send_file(file_buf, as_attachment=True, download_name=filename, mimetype='application/pdf')

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=8000)
